"""
POSH (Picture of Subsidized Households) Analysis
Generates posh_results.json from HUD POSH state and national data.

Source: HUD USER, A Picture of Subsidized Households: 2024
  - US_2024_2020census.xlsx (national-level program summaries)
  - STATE_2024_2020census.xlsx (state-level data)
  - dictionary_2024.pdf (data dictionary)

See Appendix L, Section L.2 for methodology documentation.
"""

import openpyxl
import json
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(REPO_DIR, "data")
RESULTS_DIR = os.path.join(REPO_DIR, "results")


def safe_int(val):
    if val is None or val == "" or val == ".":
        return 0
    try:
        return int(float(str(val).replace(",", "")))
    except (ValueError, TypeError):
        return 0


def safe_pct(val):
    if val is None or val == "" or val == "." or val == -1:
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0


def load_national_data():
    """Load national-level POSH data from US_2024_2020census.xlsx."""
    path = os.path.join(DATA_DIR, "US_2024_2020census.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    def col_idx(name):
        for i, h in enumerate(headers):
            if name.lower() in h.lower():
                return i
        return None

    programs = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        program_name = str(row_list[0]).strip() if row_list[0] else ""
        if not program_name:
            continue

        total_units = safe_int(row_list[col_idx("total_units")] if col_idx("total_units") is not None else 0)
        number_reported = safe_int(row_list[col_idx("number_reported")] if col_idx("number_reported") is not None else 0)

        pct_dis_lt62 = safe_pct(row_list[col_idx("pct_disabled_lt62")] if col_idx("pct_disabled_lt62") is not None else 0)
        pct_dis_ge62 = safe_pct(row_list[col_idx("pct_disabled_ge62")] if col_idx("pct_disabled_ge62") is not None else 0)
        pct_dis_all = safe_pct(row_list[col_idx("pct_disabled_all")] if col_idx("pct_disabled_all") is not None else 0)
        pct_age62plus = safe_pct(row_list[col_idx("pct_age62plus")] if col_idx("pct_age62plus") is not None else 0)

        # Weighted disability rate: combine under-62 and 62+ rates
        share_62plus = pct_age62plus / 100 if pct_age62plus else 0
        share_under62 = 1 - share_62plus
        weighted_rate = round(pct_dis_lt62 * share_under62 + pct_dis_ge62 * share_62plus, 1)

        est_disabled_hh = round(number_reported * weighted_rate / 100)

        programs[program_name] = {
            "total_units": total_units,
            "number_reported": number_reported,
            "pct_disabled_lt62": pct_dis_lt62,
            "pct_disabled_ge62": pct_dis_ge62,
            "pct_disabled_all": pct_dis_all,
            "pct_age62plus": pct_age62plus,
            "est_disabled_hh": est_disabled_hh,
            "weighted_disability_rate": weighted_rate,
        }

    wb.close()
    return programs


def load_state_data():
    """Load state-level POSH data from STATE_2024_2020census.xlsx."""
    path = os.path.join(DATA_DIR, "STATE_2024_2020census.xlsx")
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else "" for cell in ws[1]]

    def col_idx(name):
        for i, h in enumerate(headers):
            if name.lower() in h.lower():
                return i
        return None

    states = {}
    for row in ws.iter_rows(min_row=2, values_only=True):
        row_list = list(row)
        state_name = str(row_list[0]).strip() if row_list[0] else ""
        if not state_name:
            continue

        total_units = safe_int(row_list[col_idx("total_units")] if col_idx("total_units") is not None else 0)
        number_reported = safe_int(row_list[col_idx("number_reported")] if col_idx("number_reported") is not None else 0)
        pct_dis_lt62 = safe_pct(row_list[col_idx("pct_disabled_lt62")] if col_idx("pct_disabled_lt62") is not None else 0)
        pct_dis_ge62 = safe_pct(row_list[col_idx("pct_disabled_ge62")] if col_idx("pct_disabled_ge62") is not None else 0)
        pct_dis_all = safe_pct(row_list[col_idx("pct_disabled_all")] if col_idx("pct_disabled_all") is not None else 0)

        states[state_name] = {
            "total_units": total_units,
            "number_reported": number_reported,
            "pct_disabled_lt62": pct_dis_lt62,
            "pct_disabled_ge62": pct_dis_ge62,
            "pct_disabled_all": pct_dis_all,
        }

    wb.close()
    return states


def main():
    print("Loading POSH national data...")
    programs = load_national_data()

    print("Loading POSH state data...")
    states = load_state_data()

    # Build summary
    summary_program = programs.get("Summary of All HUD Programs", {})
    total_units = summary_program.get("total_units", 0)
    total_reported = summary_program.get("number_reported", 0)

    results = {
        "metadata": {
            "source": "HUD Picture of Subsidized Households 2024",
            "url": "https://www.huduser.gov/portal/datasets/assthsg.html",
            "data_dictionary": "https://www.huduser.gov/portal/datasets/pictures/dictionary_2024.pdf",
            "snapshot_date": "2024-12-31",
            "analysis_date": "2026-04-12",
        },
        "national_summary": {
            "total_subsidized_units": total_units,
            "total_reported_households": total_reported,
            "disability_rates": {
                "pct_disabled_head_spouse_cohead_under_62": summary_program.get("pct_disabled_lt62", 0),
                "pct_disabled_head_spouse_cohead_62_plus": summary_program.get("pct_disabled_ge62", 0),
                "pct_all_persons_with_disability": summary_program.get("pct_disabled_all", 0),
                "weighted_household_disability_rate": summary_program.get("weighted_disability_rate", 0),
            },
            "estimated_disabled_households": {
                "total": summary_program.get("est_disabled_hh", 0),
            },
        },
        "by_program": programs,
        "state_count": len(states),
    }

    output_path = os.path.join(RESULTS_DIR, "posh_results.json")
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2)

    print(f"Results saved to {output_path}")
    print(f"\nNational summary:")
    print(f"  Total subsidized units: {total_units:,}")
    print(f"  Weighted disability rate: {summary_program.get('weighted_disability_rate', 0)}%")
    print(f"  Estimated disabled households: ~{summary_program.get('est_disabled_hh', 0):,}")
    print(f"\nProgram breakdown:")
    for name, data in programs.items():
        if name != "Summary of All HUD Programs" and data.get("est_disabled_hh", 0) > 0:
            print(f"  {name}: {data['est_disabled_hh']:,} disabled HH ({data['weighted_disability_rate']}%)")


if __name__ == "__main__":
    main()
