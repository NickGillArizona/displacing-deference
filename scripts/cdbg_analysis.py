import openpyxl
import json
import os

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(REPO_DIR, "data")
RESULTS_DIR = os.path.join(REPO_DIR, "results")
wb = openpyxl.load_workbook(os.path.join(DATA_DIR, "CDBG_Accomp_Natl.xlsx"), data_only=True)
ws = wb["National Accomplishments"]

all_rows = []
for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
    all_rows.append(list(row))

# FY column mapping (benefitting columns, regular CDBG only)
fy_cols = {
    2024: 2, 2023: 6, 2022: 10, 2021: 14,
    2020: 18, 2019: 20, 2018: 22, 2017: 24,
    2016: 26, 2015: 28, 2014: 30, 2013: 32,
    2012: 34, 2011: 36, 2010: 38, 2009: 40,
    2008: 42, 2007: 44, 2006: 46, 2005: 48
}

# CDBG Formula Grant appropriations (billions)
cdbg_formula_grants = {
    2005: 4.117, 2006: 3.711, 2007: 3.711, 2008: 3.593,
    2009: 3.642, 2010: 3.948, 2011: 3.303, 2012: 2.948,
    2013: 3.078, 2014: 3.030, 2015: 3.066, 2016: 3.000,
    2017: 3.000, 2018: 3.365, 2019: 3.365, 2020: 3.425,
    2021: 3.450, 2022: 3.300, 2023: 3.300, 2024: 3.300
}

def safe_float(val):
    if val is None or val == "":
        return 0
    try:
        return float(val)
    except (ValueError, TypeError):
        return 0

def get_yearly(row):
    return {yr: safe_float(row[col]) for yr, col in sorted(fy_cols.items())}


# PUBLIC SERVICES (rows 49-71, 0-indexed 48-70)
ps_data = {}
for row_idx in range(48, 71):
    row = all_rows[row_idx]
    code = str(row[0]).strip() if row[0] else ""
    name = str(row[1]).strip() if row[1] else ""
    if not code or code == "Matrix CD":
        continue
    ps_data[code] = {"name": name, "yearly": get_yearly(row)}

ps_total_benefitting = get_yearly(all_rows[71])

# PUBLIC IMPROVEMENTS (rows 83-102, 0-indexed 82-101)
pi_data = {}
for row_idx in range(82, 102):
    row = all_rows[row_idx]
    code = str(row[0]).strip() if row[0] else ""
    name = str(row[1]).strip() if row[1] else ""
    if not code or code == "Matrix CD":
        continue
    pi_data[code] = {"name": name, "yearly": get_yearly(row)}

pi_total = get_yearly(all_rows[102])

# HOUSING (rows 5-18, 0-indexed 4-17)
housing_data = {}
for row_idx in range(4, 18):
    row = all_rows[row_idx]
    code = str(row[0]).strip() if row[0] else ""
    name = str(row[1]).strip() if row[1] else ""
    if not code or code == "Matrix CD":
        continue
    housing_data[code] = {"name": name, "yearly": get_yearly(row)}

housing_total = get_yearly(all_rows[18])

# Key disability categories
d05b = ps_data.get("05B", {}).get("yearly", {})
d03b = pi_data.get("03B", {}).get("yearly", {})
d05j = ps_data.get("05J", {}).get("yearly", {})

# Build results
results = {
    "source": "HUD CDBG National Accomplishments Report (CDBG_Accomp_Natl.xlsx)",
    "source_url": "https://files.hudexchange.info/resources/documents/CDBG_Accomp_Natl.xlsx",
    "data_type": "Persons/households served (accomplishments), NOT dollar expenditures",
    "fiscal_years_covered": "FY2005-FY2024 (20 years)",
    "notes": [
        "Data measures persons benefitting from CDBG activities, not dollar spending",
        "CDBG grantees must certify they will affirmatively further fair housing (AFFH) under 42 USC 5304(b)(2)",
        "Matrix codes 05B and 03B are the only disability-specific CDBG categories",
        "No separate category exists for ADA compliance, housing accessibility, or disability-related housing modifications",
        "Housing rehabilitation (14A-14J) may include some accessibility modifications but these are not tracked separately"
    ],
    "cdbg_formula_grant_appropriations_billions": cdbg_formula_grants,
    "activity_categories": {
        "housing": {code: {"name": d["name"]} for code, d in housing_data.items()},
        "public_services": {code: {"name": d["name"]} for code, d in ps_data.items()},
        "public_improvements": {code: {"name": d["name"]} for code, d in pi_data.items()},
    }
}

# Year-by-year disability analysis
disability_analysis = {}
for yr in sorted(fy_cols.keys()):
    ps_ben_total = ps_total_benefitting.get(yr, 0)
    pi_ben_total = pi_total.get(yr, 0)
    h_total = housing_total.get(yr, 0)

    services_disability = d05b.get(yr, 0)
    facilities_disability = d03b.get(yr, 0)
    combined_disability = services_disability + facilities_disability
    fair_housing = d05j.get(yr, 0)

    ps_pct = (services_disability / ps_ben_total * 100) if ps_ben_total > 0 else None
    pi_pct = (facilities_disability / pi_ben_total * 100) if pi_ben_total > 0 else None

    disability_analysis[yr] = {
        "05B_disability_services_persons": int(services_disability),
        "03B_disability_facilities_persons": int(facilities_disability),
        "combined_disability_persons": int(combined_disability),
        "05J_fair_housing_persons": int(fair_housing),
        "public_services_total_benefitting": int(ps_ben_total),
        "public_improvements_total_benefitting": int(pi_ben_total),
        "housing_total_households": int(h_total),
        "disability_pct_of_public_services": round(ps_pct, 3) if ps_pct else None,
        "disability_pct_of_public_improvements": round(pi_pct, 3) if pi_pct else None,
        "cdbg_formula_grant_billions": cdbg_formula_grants.get(yr)
    }

results["disability_analysis_by_year"] = disability_analysis

# Summary statistics
total_05b = sum(d05b.values())
total_03b = sum(d03b.values())
total_05j = sum(d05j.values())
total_ps = sum(ps_total_benefitting.values())
total_pi = sum(pi_total.values())

avg_05b = total_05b / 20
avg_03b = total_03b / 20

early_03b = sum(d03b.get(yr, 0) for yr in range(2005, 2010)) / 5
late_03b = sum(d03b.get(yr, 0) for yr in range(2019, 2024)) / 5
early_05b = sum(d05b.get(yr, 0) for yr in range(2005, 2010)) / 5
late_05b = sum(d05b.get(yr, 0) for yr in range(2019, 2024)) / 5

results["summary_statistics"] = {
    "twenty_year_totals": {
        "disability_services_05B_total_persons": int(total_05b),
        "disability_facilities_03B_total_persons": int(total_03b),
        "combined_disability_total_persons": int(total_05b + total_03b),
        "fair_housing_05J_total_persons": int(total_05j),
        "public_services_total_benefitting": int(total_ps),
        "public_improvements_total_benefitting": int(total_pi),
    },
    "twenty_year_averages": {
        "avg_annual_disability_services_05B": round(avg_05b),
        "avg_annual_disability_facilities_03B": round(avg_03b),
        "avg_annual_public_services_total": round(total_ps / 20),
        "avg_annual_public_improvements_total": round(total_pi / 20),
    },
    "share_calculations": {
        "05B_share_of_public_services_20yr_pct": round(total_05b / total_ps * 100, 3) if total_ps > 0 else None,
        "03B_share_of_public_improvements_20yr_pct": round(total_03b / total_pi * 100, 3) if total_pi > 0 else None,
    },
    "trend_03B_facilities": {
        "avg_FY2005_2009": round(early_03b),
        "avg_FY2019_2023": round(late_03b),
        "change_pct": round((late_03b - early_03b) / early_03b * 100, 1) if early_03b > 0 else None
    },
    "trend_05B_services": {
        "avg_FY2005_2009": round(early_05b),
        "avg_FY2019_2023": round(late_05b),
        "change_pct": round((late_05b - early_05b) / early_05b * 100, 1) if early_05b > 0 else None
    },
    "total_cdbg_formula_grants_20yr_billions": round(sum(cdbg_formula_grants.values()), 3)
}

results["argument_support"] = {
    "finding_1": (
        "HUD CDBG accomplishment data contains exactly TWO disability-specific activity "
        "categories out of approximately 50 total: 05B (Services for Persons with "
        "Disabilities) and 03B (Facility for Persons with Disabilities)."
    ),
    "finding_2": (
        f"Over 20 years (FY2005-FY2024), disability services (05B) served "
        f"{int(total_05b):,} persons - only "
        f"{round(total_05b / total_ps * 100, 2)}% of all "
        f"{int(total_ps):,} public service beneficiaries."
    ),
    "finding_3": (
        f"Disability facilities (03B) served {int(total_03b):,} persons - only "
        f"{round(total_03b / total_pi * 100, 2)}% of all "
        f"{int(total_pi):,} public improvement beneficiaries."
    ),
    "finding_4": (
        "NO separate CDBG category exists for ADA compliance, housing accessibility "
        "modifications, or disability-related housing rehabilitation. Any accessibility "
        "work is buried in general rehabilitation codes (14A-14J) with no disability-specific tracking."
    ),
    "finding_5": (
        f"Fair Housing Activities (05J) served {int(total_05j):,} persons over 20 years, "
        "but there is no subcategory for disability-specific fair housing work, despite "
        "disability being a protected class under the Fair Housing Act."
    ),
    "finding_6": (
        f"CDBG grantees received ${round(sum(cdbg_formula_grants.values()), 1)} billion "
        "in formula grants over this period while certifying they would affirmatively "
        "further fair housing, yet disability-specific accomplishments represent a "
        "negligible fraction of total activity."
    ),
    "finding_7": (
        "The absence of disability-specific tracking categories for housing accessibility "
        "is itself evidence of performative AFFH certification - grantees cannot "
        "meaningfully report on disability accessibility because the reporting system "
        "does not require it."
    ),
    "finding_8": (
        f"Disability facility construction (03B) declined from an average of "
        f"{int(early_03b):,} persons/year (FY2005-2009) to "
        f"{int(late_03b):,} persons/year (FY2019-2023), a "
        f"{round((late_03b - early_03b) / early_03b * 100, 1)}% change."
    )
}

# Save
output_path = os.path.join(RESULTS_DIR, "cdbg_results.json")
with open(output_path, "w") as f:
    json.dump(results, f, indent=2)

print(f"Results saved to {output_path}")
print(f"File size: {os.path.getsize(output_path):,} bytes")

# Print summary
print("\n" + "=" * 80)
print("SUMMARY OF KEY FINDINGS")
print("=" * 80)
for k, v in results["argument_support"].items():
    print(f"\n{k}: {v}")

print("\n" + "=" * 80)
print("YEAR-BY-YEAR DISABILITY SHARE")
print("=" * 80)
hdr = f"{'Year':>6} | {'05B Served':>12} | {'05B % of PS':>12} | {'03B Served':>12} | {'03B % of PI':>12} | {'CDBG $B':>8}"
print(hdr)
print("-" * len(hdr))
for yr in sorted(disability_analysis.keys()):
    d = disability_analysis[yr]
    ps_str = f"{d['disability_pct_of_public_services']:.2f}%" if d["disability_pct_of_public_services"] else "N/A"
    pi_str = f"{d['disability_pct_of_public_improvements']:.2f}%" if d["disability_pct_of_public_improvements"] else "N/A"
    print(
        f"{yr:>6} | {d['05B_disability_services_persons']:>12,} | "
        f"{ps_str:>12} | {d['03B_disability_facilities_persons']:>12,} | "
        f"{pi_str:>12} | {d['cdbg_formula_grant_billions']:>8.3f}"
    )

wb.close()
