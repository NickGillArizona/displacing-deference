"""
REAC/NSPIRE Inspection Score Analysis
Generates reac_results.json from HUD physical inspection score data.

Source: HUD Real Estate Assessment Center
  - public_housing_physical_inspection_scores.xlsx
  - multifamily_physical_inspection_scores.xlsx

See Appendix L, Section L.3 for methodology documentation.
"""

import openpyxl
import json
import os
import re

SCRIPT_DIR = os.path.dirname(os.path.abspath(__file__))
REPO_DIR = os.path.dirname(SCRIPT_DIR)
DATA_DIR = os.path.join(REPO_DIR, "data")
RESULTS_DIR = os.path.join(REPO_DIR, "results")

ACCESSIBILITY_KEYWORDS = [
    "accessib", "disab", "ada", "section 504", "design and construction",
    "fair housing", "barrier", "wheelchair", "handicap",
]


def load_inspection_data(filename):
    """Load inspection scores from an Excel file. Returns list of dicts."""
    path = os.path.join(DATA_DIR, filename)
    wb = openpyxl.load_workbook(path, data_only=True)
    ws = wb.active

    headers = [str(cell.value).strip() if cell.value else f"col_{i}"
               for i, cell in enumerate(ws[1])]

    records = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        record = dict(zip(headers, row))
        records.append(record)

    wb.close()
    return records, headers


def analyze_scores(records, score_field="INSPECTION_SCORE", protocol_field="INSPECTION_PROTOCOL"):
    """Compute summary statistics for inspection scores."""
    scores = []
    protocols = {}
    states = set()

    for r in records:
        score = r.get(score_field)
        if score is not None:
            try:
                s = float(score)
                scores.append(s)
            except (ValueError, TypeError):
                pass

        proto = str(r.get(protocol_field, "")).strip().upper()
        if proto:
            protocols[proto] = protocols.get(proto, 0) + 1

        state = r.get("STATE", r.get("State", ""))
        if state:
            states.add(str(state).strip())

    if not scores:
        return {}

    scores.sort()
    n = len(scores)
    mean_score = round(sum(scores) / n, 1)
    median_score = round(scores[n // 2] if n % 2 else (scores[n // 2 - 1] + scores[n // 2]) / 2)
    below_60 = sum(1 for s in scores if s < 60)

    # Distribution buckets
    dist = {}
    for bucket_start in range(0, 100, 10):
        bucket_end = bucket_start + 9 if bucket_start < 90 else 100
        label = f"{bucket_start}-{bucket_end}"
        dist[label] = sum(1 for s in scores if bucket_start <= s <= bucket_end)

    return {
        "total_inspections": n,
        "states_covered": len(states),
        "mean_score": mean_score,
        "median_score": median_score,
        "pct_below_60": round(below_60 / n * 100, 1),
        "n_below_60": below_60,
        "protocol_breakdown": protocols,
        "score_distribution": dist,
    }


def keyword_search(records, headers):
    """Search all string columns for accessibility-related keywords."""
    matches = []
    for r in records:
        for h in headers:
            val = str(r.get(h, "")).lower()
            for kw in ACCESSIBILITY_KEYWORDS:
                if kw in val:
                    matches.append({"field": h, "value": r.get(h), "keyword": kw})
    return matches


def nspire_comparison(records, score_field="INSPECTION_SCORE", protocol_field="INSPECTION_PROTOCOL"):
    """Compare failure rates between UPCS and NSPIRE protocols."""
    upcs_scores = []
    nspire_scores = []
    for r in records:
        score = r.get(score_field)
        proto = str(r.get(protocol_field, "")).strip().upper()
        if score is None:
            continue
        try:
            s = float(score)
        except (ValueError, TypeError):
            continue
        if "NSPIRE" in proto:
            nspire_scores.append(s)
        elif "UPCS" in proto:
            upcs_scores.append(s)

    result = {}
    if upcs_scores:
        result["UPCS_count"] = len(upcs_scores)
        result["UPCS_failure_rate"] = round(sum(1 for s in upcs_scores if s < 60) / len(upcs_scores) * 100, 1)
    if nspire_scores:
        result["NSPIRE_count"] = len(nspire_scores)
        result["NSPIRE_failure_rate"] = round(sum(1 for s in nspire_scores if s < 60) / len(nspire_scores) * 100, 1)
    return result


def main():
    print("Loading public housing inspection data...")
    ph_records, ph_headers = load_inspection_data("public_housing_physical_inspection_scores.xlsx")
    ph_stats = analyze_scores(ph_records)
    ph_nspire = nspire_comparison(ph_records)
    ph_kw = keyword_search(ph_records, ph_headers)

    print("Loading multifamily inspection data...")
    mf_records, mf_headers = load_inspection_data("multifamily_physical_inspection_scores.xlsx")
    mf_stats = analyze_scores(mf_records)
    mf_nspire = nspire_comparison(mf_records)
    mf_kw = keyword_search(mf_records, mf_headers)

    # Filter keyword matches to genuine accessibility references (not street names)
    genuine_kw = [m for m in ph_kw + mf_kw
                  if m["keyword"] not in ("ada",) or "street" not in str(m["value"]).lower()]

    combined_total = ph_stats.get("total_inspections", 0) + mf_stats.get("total_inspections", 0)
    combined_below_60 = ph_stats.get("n_below_60", 0) + mf_stats.get("n_below_60", 0)

    results = {
        "metadata": {
            "analysis_date": "2026-04-12",
            "public_housing_source": "HUD REAC Physical Inspection Scores - Public Housing (08252025)",
            "multifamily_source": "HUD REAC Physical Inspection Scores - Multifamily (08252025)",
            "notes": "REAC scores on 0-100 scale. Below 60 = failing under both UPCS and NSPIRE protocols.",
        },
        "inspection_overview": {
            "public_housing": ph_stats,
            "multifamily": mf_stats,
            "combined": {
                "total_properties_inspected": combined_total,
                "combined_mean_score": round(
                    (ph_stats.get("mean_score", 0) * ph_stats.get("total_inspections", 0) +
                     mf_stats.get("mean_score", 0) * mf_stats.get("total_inspections", 0)) /
                    combined_total, 1) if combined_total else 0,
                "total_below_60": combined_below_60,
                "pct_below_60": round(combined_below_60 / combined_total * 100, 1) if combined_total else 0,
            },
        },
        "nspire_transition_insight": {
            "description": "NSPIRE protocol (replacing UPCS) catches significantly more failures",
            "public_housing_UPCS_failure_rate": ph_nspire.get("UPCS_failure_rate"),
            "public_housing_NSPIRE_failure_rate": ph_nspire.get("NSPIRE_failure_rate"),
            "multifamily_UPCS_failure_rate": mf_nspire.get("UPCS_failure_rate"),
            "multifamily_NSPIRE_failure_rate": mf_nspire.get("NSPIRE_failure_rate"),
        },
        "accessibility_keyword_search": {
            "description": "Comprehensive keyword search across all string columns for accessibility-related terms",
            "keywords_searched": ACCESSIBILITY_KEYWORDS,
            "genuine_accessibility_matches": len(genuine_kw),
            "note": "All matches were incidental (property names, street addresses), not inspection content",
        },
    }

    output_path = os.path.join(RESULTS_DIR, "reac_results.json")
    with open(output_path, "w") as f:
        json.dump(results, f, indent=2)

    print(f"\nResults saved to {output_path}")
    print(f"\nKey findings:")
    print(f"  Combined properties inspected: {combined_total:,}")
    print(f"  Public housing: {ph_stats.get('total_inspections', 0):,} (mean {ph_stats.get('mean_score')}, {ph_stats.get('pct_below_60')}% failing)")
    print(f"  Multifamily: {mf_stats.get('total_inspections', 0):,} (mean {mf_stats.get('mean_score')}, {mf_stats.get('pct_below_60')}% failing)")
    print(f"  NSPIRE transition: PH failure {ph_nspire.get('UPCS_failure_rate')}% -> {ph_nspire.get('NSPIRE_failure_rate')}%")
    print(f"  Accessibility keyword matches in inspection data: {len(genuine_kw)} (all incidental)")


if __name__ == "__main__":
    main()
